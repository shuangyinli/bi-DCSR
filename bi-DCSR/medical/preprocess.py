'''
Created on Dec 17, 2019

@author: shuangyinli
'''

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter 
import random
from numpy import string_
import numpy as np

workbook_ = load_workbook("/Users/huihui/Data/medical/data/mtsamples.xlsx")

outfile = open("/Users/huihui/Data/medical/process/plain.txt", "w")
labeloutfile = open("/Users/huihui/Data/medical/process/labels.txt","w")
sheetnames =workbook_.get_sheet_names() 
sheetr = workbook_.get_sheet_by_name(sheetnames[0])

labellist = []
mainlist = []

labels = ["Cardiovascular / Pulmonary", "Gastroenterology", "General Medicine", "Neurology", "Obstetrics / Gynecology", "Orthopedic", "Radiology", "Surgery", "Urology"]

for i in range(sheetr.max_row):
    label = sheetr.cell(row=i+1,column=3).value.lstrip().rstrip()
    mainquestion = sheetr.cell(row=i+1,column=5).value
    if mainquestion != "" and mainquestion != " "  and label != "" and label != " " and mainquestion is not None:
        if label in labels:
            print(label)
            outfile.write(label.lstrip().rstrip() + "##")
            labeloutfile.write(label.lstrip().rstrip() +"\n")
            outfile.write(mainquestion.strip().rstrip() + "\n")
    pass

outfile.close()
labeloutfile.close()

