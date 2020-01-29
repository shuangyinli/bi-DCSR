'''
Created on May 31, 2019

@author: shuangyinli
'''
# coding: utf-8
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter 
import jieba
import random
from numpy import string_

template = [u"具体高校信息请到院校详情页点击“问招办”查询。",u"详细请咨询高校招办"]
template2 = [u"您的提问小i还未学习到，请去院校详情页查看吧"]

def update(a):
    if a in template:
        return u"请前往m.wmzy.com下载完美志愿APP进行查询"
    elif a in template2:
        return u"这个问题大师兄也不知道啊"
    else:
        b = a[0:588]
        return b.replace("您好，", "你")

def similarlen(a):
    return a[0:60]


stopwords = {}
def readstopwords(stopwordsfile):
    return {}.fromkeys([ line.rstrip() for line in open(stopwordsfile, 'r', encoding='utf-8') ])

stopwords = readstopwords("cn_stopwords.txt")

workbook_ = load_workbook("/Users/huihui/Downloads/export_446759913_20190531183906.xlsx")

sheetnames =workbook_.get_sheet_names() 
sheetr = workbook_.get_sheet_by_name(sheetnames[0])

finallist = []
index = False

indexQ = ""

onequestionanswerlist = []
onesimilarQ = []

for i in range(sheetr.max_row):
    mainquestion = sheetr.cell(row=i+1,column=1).value
    if mainquestion !=None:
        finallist.append((onequestionanswerlist,onesimilarQ))
        index =True
        indexQ = str(sheetr.cell(row=i+1,column=5).value).rstrip().lstrip()
        #do
        onequestionanswerlist = []
        onesimilarQ = []
        onequestionanswerlist.append((indexQ,str(sheetr.cell(row=i+1,column=8).value).rstrip().lstrip()))
        onesimilarQ.append(str(sheetr.cell(row=i+1,column=11).value).rstrip().lstrip())
    else:
        index =False
    
    if index == False:
        onequestionanswerlist.append((str(sheetr.cell(row=i+1,column=6).value).rstrip().lstrip() + "#" + indexQ, str(sheetr.cell(row=i+1,column=8).value).rstrip().lstrip()))
        onesimilarQ.append(str(sheetr.cell(row=i+1,column=11).value).rstrip().lstrip() )
        pass
    
wb = Workbook()
sheetw = wb.active

for item in finallist:
    onesimilarQlist = item[1]
    similarQueslist =[]
    similarQueslistofset = set()
    for similarquestion in onesimilarQlist:
        if similarquestion != " " and similarquestion != "" and similarquestion != "None" and similarquestion != None:
            seg_list = jieba.cut(similarquestion, cut_all=True)
            seg_list_temp = [it for it in seg_list if it not in stopwords]
            string_seg_list_temp = " ".join(seg_list_temp)
            similarQueslist.append(string_seg_list_temp)
    
        #TODO more intellegent
        #similarQueslistofset.add(list(set(seg_list_temp)))
        #if set(seg_list_temp) not in similarQueslistofset:
            #keep the sequence info
            #similarQueslist.append(" ".join(seg_list_temp))
    
    random.shuffle(similarQueslist)
    #if len(similarQueslist) > 10:
        #similarQueslist = similarQueslist[:10]
    
    onequestionanswerlist = item[0]
    if len(onequestionanswerlist) <= 0:
        continue
    q1 = onequestionanswerlist[0][0]
    a1 = onequestionanswerlist[0][1]
    
    if a1 != " " and a1!="":
        finalwriteqa = []
        finalwriteqa.append(q1)
        finalwriteqa.append(update(a1))
        finalwriteqa += similarQueslist
        sheetw.append(finalwriteqa)
    
    for qa in onequestionanswerlist[1:]:
        q = qa[0]
        a = qa[1]
        if a != " " and a !="" and a != "None" and a != None:
            qlist = q.rstrip().lstrip().split("#")
            schoolname = qlist[0]
            mainq = qlist[1]
            finalwriteq = []
            finalwriteq.append(schoolname + " " +mainq)
            finalwriteq.append(update(a))
        
            for item in similarQueslist:
                finalwriteq.append(similarlen(schoolname + " " + item))
                
            sheetw.append(finalwriteq)
 
wb.save("/Users/huihui/Downloads/test.csv")
        
        
        
        






